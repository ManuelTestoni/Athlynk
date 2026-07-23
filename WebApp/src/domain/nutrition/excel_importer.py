"""Pipeline import dieta da Excel.

Flusso API-first puro (input → output, niente Django request/response):
  1. excel_to_text: bytes → stringa griglia testuale
  2. extract_diet_with_ai: stringa → dict JSON (chiamata LLM, response_format json)
  3. normalize_and_match: dict raw → dict con food_id e candidati Food
  4. compute_confidence: dict normalizzato → summary uncertain
"""

import io
import json
import re

from openpyxl import load_workbook
from langchain_core.messages import SystemMessage, HumanMessage
from pydantic import ValidationError

from domain.nutrition.food_match import best_match
from domain.nutrition.llm_client import build_extraction_llm
from domain.nutrition.schemas import DietExtraction, ConfidenceSummary
from domain.shared.doc_structure import DocStructure, detect_diet_structure
from domain.shared.excel_text import ExcelParseError  # noqa: F401  (re-export)
from domain.shared.extraction import AIExtractionError  # noqa: F401  (re-export)


MAX_GRID_CHARS = 12000


SYSTEM_PROMPT = """Sei un estrattore di dati nutrizionali. Ricevi il contenuto grezzo di un file Excel
creato da un nutrizionista e devi mapparlo nello schema JSON standard.

REGOLE:
- Restituisci SOLO JSON valido, nessun testo fuori dal JSON
- Per i campi non trovati usa null, MAI inventare dati
- Normalizza i giorni in inglese maiuscolo: MONDAY, TUESDAY, WEDNESDAY, THURSDAY, FRIDAY, SATURDAY, SUNDAY
- Normalizza i pasti: BREAKFAST, MORNING_SNACK, LUNCH, AFTERNOON_SNACK, DINNER
- Normalizza le unità: g, ml, portion, tbsp, tsp
- Per campi ambigui aggiungi "uncertain": true accanto al valore
- Dizionario sinonimi italiani da gestire:
  Giorni: lunedì/lun/giorno1 → MONDAY, martedì/mar → TUESDAY, mercoledì/mer → WEDNESDAY,
          giovedì/gio → THURSDAY, venerdì/ven → FRIDAY, sabato/sab → SATURDAY, domenica/dom → SUNDAY
  Pasti: colazione/col → BREAKFAST, spuntino_mattina/spunt.matt → MORNING_SNACK,
         pranzo/prz → LUNCH, merenda/spuntino_pomeriggio → AFTERNOON_SNACK, cena → DINNER
  Nutrienti: kcal/calorie/cal → calories, proteine/prot/PRO → protein_g,
             carboidrati/carbo/CHO/carboidr. → carbs_g, grassi/lip/FAT/lipidi → fat_g
  Unità: grammi/gr/g → g, millilitri/ml → ml, porzione/pz/pcs → portion,
         cucchiaio/cucch → tbsp, cucchiaino → tsp
- Se trovi una struttura a tabella con giorni come colonne e pasti come righe (o viceversa),
  gestiscila correttamente trasposta
- ALTERNATIVE: se un alimento è marcato come alternativa di un altro ("alternativa",
  "in alternativa", "oppure", "sostituibile con", "/"), inseriscilo nell'array
  "substitutions" dell'alimento PRINCIPALE — NON come food separato.
  Modalità: "iso-kcal"/"isocalorica"→ISOKCAL, "iso-prot"/"isoproteica"→ISOPROT,
  "iso-carb"/"isoglucidica"→ISOCARB. Default ISOKCAL.
- INTEGRATORI: se trovi righe/sezioni con "integratore"/"integratori"/"supplementi",
  estrai gli item nell'array TOP-LEVEL "supplements" con name/dose/timing/notes.
- TARGET MACRO: se il documento indica totali/target giornalieri (kcal, proteine,
  carboidrati, grassi) popolali in target_kcal/target_protein_g/target_carb_g/
  target_fat_g del giorno; se sono unici per tutto il piano usali anche nei campi
  top-level daily_kcal/protein_target_g/carb_target_g/fat_target_g.
- Includi nel JSON il campo "extraction_notes" con eventuali warning sull'estrazione

SCHEMA OUTPUT richiesto:
{
  "diet_name": string | null,
  "days": [
    {
      "day_of_week": "MONDAY",
      "target_kcal": int | null,
      "target_protein_g": int | null,
      "target_carb_g": int | null,
      "target_fat_g": int | null,
      "meals": [
        {
          "meal_type": "BREAKFAST",
          "foods": [
            {
              "name": string,
              "quantity": float | null,
              "unit": "g" | "ml" | "portion" | "tbsp" | "tsp",
              "calories": float | null,
              "protein_g": float | null,
              "carbs_g": float | null,
              "fat_g": float | null,
              "uncertain": bool,
              "notes": string | null,
              "substitutions": [
                {
                  "name": string,
                  "quantity": float | null,
                  "unit": "g"|"ml"|"portion"|"tbsp"|"tsp",
                  "mode": "ISOKCAL"|"ISOPROT"|"ISOCARB",
                  "uncertain": bool,
                  "notes": string | null
                }
              ]
            }
          ]
        }
      ]
    }
  ],
  "supplements": [
    {
      "name": string,
      "dose": string | null,
      "timing": string | null,
      "notes": string | null,
      "uncertain": bool
    }
  ],
  "total_calories_daily": float | null,
  "daily_kcal": int | null,
  "protein_target_g": int | null,
  "carb_target_g": int | null,
  "fat_target_g": int | null,
  "extraction_notes": string | null
}
"""


MACRO_SYSTEM_PROMPT = """Sei un estrattore di piani nutrizionali A SOLI MACRO (nessun alimento
specifico: solo target di kcal, proteine, carboidrati, grassi). Ricevi il testo del
documento e restituisci SOLO JSON valido:
{
  "diet_name": string | null,
  "plan_kind": "DAILY" | "WEEKLY",
  "daily_kcal": int | null,
  "protein_target_g": int | null,
  "carb_target_g": int | null,
  "fat_target_g": int | null,
  "days": [
    {"day_of_week": "MONDAY", "target_kcal": int|null, "target_protein_g": int|null,
     "target_carb_g": int|null, "target_fat_g": int|null}
  ],
  "extraction_notes": string | null
}
REGOLE:
- plan_kind=DAILY se i target sono uguali per tutti i giorni (popola i campi top-level,
  days può restare vuoto). plan_kind=WEEKLY se i target variano per giorno (popola days).
- Giorni in inglese maiuscolo MONDAY..SUNDAY (lunedì→MONDAY ecc.).
- Per valori non trovati usa null, MAI inventare.
"""


def excel_to_text(file_bytes: bytes) -> str:
    """Legge l'xlsx mantenendo la struttura visiva (coordinate cella + valore).

    Restituisce una stringa con righe del tipo:
        [SHEET: Foglio1]
        A1: Lunedì | B1: Martedì | C1: Mercoledì
        A2: Colazione | B2: Latte 200ml | ...
    Tronca a MAX_GRID_CHARS con warning.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ExcelParseError(f"File Excel non leggibile: {e}") from e

    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"[SHEET: {sheet.title}]")
        for row in sheet.iter_rows(values_only=False):
            row_parts = []
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                # Strip / collassa whitespace per evitare token-burn su celle multilinea
                text = str(val).strip().replace('\n', ' ').replace('\r', ' ')
                text = re.sub(r'\s+', ' ', text)
                if text:
                    row_parts.append(f"{cell.coordinate}: {text}")
            if row_parts:
                lines.append(' | '.join(row_parts))
        lines.append('')  # blank line between sheets

    text = '\n'.join(lines)
    if not text.strip():
        raise ExcelParseError("File Excel vuoto.")

    if len(text) > MAX_GRID_CHARS:
        text = text[:MAX_GRID_CHARS] + '\n...[CONTENUTO TRONCATO]'
    return text


def extract_diet_with_ai(
    grid_text: str,
    plan_title: str = '',
    structure: DocStructure | None = None,
) -> dict:
    """Chiama Ollama Cloud (gpt-oss) in JSON mode e ritorna dict raw.

    Riusa la stessa config di Chiron (OLLAMA_API_KEY/OLLAMA_BASE_URL).
    Piani solo-macro → prompt dedicato piccolo, una chiamata veloce.
    """
    is_macro = bool(structure and structure.layout.startswith('macro'))
    if is_macro:
        llm = build_extraction_llm(max_tokens=800, timeout=20)
        system_prompt = MACRO_SYSTEM_PROMPT
    else:
        # This is the whole-document single call (Excel grid / PDF fast path): it
        # emits the entire plan's JSON in one shot, so it needs a far larger budget
        # than a single chunk. Measured ~80s for a full weekly food grid on
        # gpt-oss:120b; the old 30s default guaranteed a timeout on food-rich diets.
        llm = build_extraction_llm(timeout=120)
        system_prompt = SYSTEM_PROMPT

    user_msg = (
        f"Titolo piano (suggerito dal nutrizionista): {plan_title or 'non specificato'}\n\n"
        f"Contenuto Excel:\n{grid_text}\n\n"
        f"Restituisci SOLO il JSON conforme allo schema."
    )

    try:
        resp = llm.invoke([
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_msg),
        ])
    except Exception as e:
        raise AIExtractionError(f"Chiamata LLM fallita: {e}") from e

    content = resp.content if isinstance(resp.content, str) else str(resp.content)
    try:
        raw = json.loads(content)
    except json.JSONDecodeError as e:
        raise AIExtractionError(f"JSON LLM invalido: {e}\nOutput: {content[:500]}") from e

    return raw


def normalize_and_match(raw_json: dict, progress_cb=None) -> dict:
    """Valida con pydantic e arricchisce ogni food con food_id + candidati.

    Per ogni food.name → fuzzy_match → se score alto: food_id+name reale,
    altrimenti uncertain=True + candidati per la UI di revisione.

    `progress_cb(done, total)`, se dato, viene chiamato dopo ogni food — pure
    query DB senza altro segnale di avanzamento durante questa fase; un job
    lungo resta "vivo" in cache solo se qualcosa la tocca (vedi JobStore ttl).
    """
    # Pydantic tollera campi extra e default; valida lo scheletro.
    try:
        validated = DietExtraction.model_validate(raw_json)
    except ValidationError as e:
        # Fallback: prova a sanare struttura minima
        raw_json.setdefault('days', [])
        validated = DietExtraction.model_validate(raw_json)

    out = validated.model_dump()

    def _match_food_into(entry: dict) -> None:
        name = (entry.get('name') or '').strip()
        if not name:
            entry['uncertain'] = True
            entry['candidates'] = []
            return
        best, others, _ = best_match(name, threshold=0.5)
        if best:
            entry['food_id'] = best['id']
            entry['matched_name'] = best['name']
            entry['candidates'] = others
            entry['db_macros'] = {
                'kcal': best.get('kcal'),
                'protein': best.get('protein'),
                'carb': best.get('carb'),
                'fat': best.get('fat'),
            }
            entry['uncertain'] = entry.get('uncertain', False)
        else:
            entry['food_id'] = None
            entry['candidates'] = others
            entry['uncertain'] = True

    all_foods = [
        food
        for day in out.get('days', [])
        for meal in day.get('meals', [])
        for food in meal.get('foods', [])
    ]
    total = len(all_foods) or 1
    for done, food in enumerate(all_foods, start=1):
        _match_food_into(food)
        for sub in food.get('substitutions', []) or []:
            _match_food_into(sub)
        if progress_cb:
            try:
                progress_cb(done, total)
            except Exception:
                pass

    # Dedup integratori per nome normalizzato, mergendo timing/notes
    def _merge_str(a: str | None, b: str | None) -> str | None:
        a = (a or '').strip(); b = (b or '').strip()
        if not b:
            return a or None
        if not a:
            return b
        if b.lower() in a.lower():
            return a
        if a.lower() in b.lower():
            return b
        return f"{a} · {b}"

    raw_supps = out.get('supplements') or []
    dedup_map: dict[str, dict] = {}
    order: list[str] = []
    for s in raw_supps:
        nkey = re.sub(r'\s+', ' ', (s.get('name') or '').strip().lower())
        if not nkey:
            continue
        if nkey in dedup_map:
            cur = dedup_map[nkey]
            cur['timing'] = _merge_str(cur.get('timing'), s.get('timing'))
            cur['notes'] = _merge_str(cur.get('notes'), s.get('notes'))
            if not cur.get('dose') and s.get('dose'):
                cur['dose'] = s.get('dose')
            continue
        dedup_map[nkey] = dict(s)
        order.append(nkey)
    out['supplements'] = [dedup_map[k] for k in order]

    # Integratori free-text: nessun catalogo, si tiene solo name/dose/timing/notes.
    for supp in out.get('supplements', []) or []:
        supp['uncertain'] = not (supp.get('name') or '').strip()

    return out


def compute_confidence(normalized: dict) -> ConfidenceSummary:
    """Conta totale food/sub/supp e quanti uncertain."""
    total = 0
    uncertain = 0
    for day in normalized.get('days', []):
        for meal in day.get('meals', []):
            for food in meal.get('foods', []):
                total += 1
                if food.get('uncertain'):
                    uncertain += 1
                for sub in food.get('substitutions', []) or []:
                    total += 1
                    if sub.get('uncertain'):
                        uncertain += 1
    for supp in normalized.get('supplements', []) or []:
        total += 1
        if supp.get('uncertain'):
            uncertain += 1
    ratio = (uncertain / total) if total else 0.0
    return ConfidenceSummary(
        fields_total=total,
        fields_uncertain=uncertain,
        ratio=round(ratio, 3),
    )


def apply_structure_hints(normalized: dict, structure: DocStructure | None) -> dict:
    """Attach plan_mode/plan_kind hints (LLM value wins, detector fills gaps)."""
    if structure is None:
        return normalized
    if not normalized.get('plan_mode'):
        normalized['plan_mode'] = 'MACRO' if structure.layout.startswith('macro') else 'FOOD'
    if not normalized.get('plan_kind'):
        normalized['plan_kind'] = 'WEEKLY' if structure.layout.endswith('weekly') else 'DAILY'
    return normalized


def run_import_pipeline(file_bytes: bytes, plan_title: str = '',
                        progress_cb=None) -> tuple[dict, ConfidenceSummary]:
    """Helper end-to-end. Solleva ExcelParseError o AIExtractionError."""
    grid = excel_to_text(file_bytes)
    structure = detect_diet_structure(grid)
    raw = extract_diet_with_ai(grid, plan_title, structure=structure)
    normalized = normalize_and_match(raw, progress_cb=progress_cb)
    apply_structure_hints(normalized, structure)
    confidence = compute_confidence(normalized)
    return normalized, confidence
