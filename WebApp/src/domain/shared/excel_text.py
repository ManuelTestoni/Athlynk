"""Excel → flat text grid. Domain-neutral.

Reads xlsx preserving cell coordinates so the LLM can reconstruct row/col
relationships. Truncates at MAX_GRID_CHARS.
"""

import io
import re

from openpyxl import load_workbook


MAX_GRID_CHARS = 12000


class ExcelParseError(Exception):
    """Excel file unreadable or empty."""


def excel_to_text(file_bytes: bytes, max_chars: int = MAX_GRID_CHARS) -> str:
    """Return a string with rows like `A1: cell | B1: cell | ...` per sheet."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ExcelParseError(f"File Excel non leggibile: {e}") from e

    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"[SHEET: {sheet.title}]")
        for row in sheet.iter_rows(values_only=False):
            row_parts = []
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                text = str(val).strip().replace('\n', ' ').replace('\r', ' ')
                text = re.sub(r'\s+', ' ', text)
                if text:
                    row_parts.append(f"{cell.coordinate}: {text}")
            if row_parts:
                lines.append(' | '.join(row_parts))
        lines.append('')

    text = '\n'.join(lines)
    if not text.strip():
        raise ExcelParseError("File Excel vuoto.")
    if len(text) > max_chars:
        text = text[:max_chars] + '\n...[CONTENUTO TRONCATO]'
    return text
