#!/usr/bin/env python
"""
Standalone importer for Exercise_DB_it.xlsx → Exercise table.

Usage:
    cd WebApp/src
    ../../venv/bin/python import_exercises.py
    # or with explicit path:
    ../../venv/bin/python import_exercises.py /absolute/path/to/Exercise_DB_it.xlsx

Behavior:
- Loads xlsx (header row 1, data from row 2).
- If Exercise table has rows, prompts for confirmation, then deletes everything
  before reinserting (cascade applies on related WorkoutExercise rows).
- Inserts each row, generating a unique slug from name.
- Prints progress every row.

Column mapping (xlsx → model field):
    Exercise                                   → name
    Video dimostrativo (hyperlink target)      → video_url
    Livello di difficoltà                      → difficulty_level
    Gruppo muscolare bersaglio                 → target_muscle_group
    Muscolo motore principale                  → primary_muscle
    Muscolo secondario                         → secondary_muscle
    Attrezzatura principale                    → equipment
    Schema di movimento n. 1                   → movement_pattern_1
    Schema di movimento n. 2                   → movement_pattern_2
    Regione corporea                           → body_region
    Classificazione primaria dell'esercizio    → exercise_classification
"""
import os
import sys
import re
import django
from pathlib import Path

# Bootstrap Django
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.utils.text import slugify
from django.db import transaction
from openpyxl import load_workbook

from domain.workouts.models import Exercise


DEFAULT_XLSX = Path(__file__).resolve().parents[2] / 'Exercise_DB_it.xlsx'

COLUMN_INDEX = {
    'name':                    0,   # Exercise
    'video_url':               1,   # Video dimostrativo (hyperlink)
    'difficulty_level':        2,   # Livello di difficoltà
    'target_muscle_group':     3,   # Gruppo muscolare bersaglio
    'primary_muscle':          4,   # Muscolo motore principale
    'secondary_muscle':        5,   # Muscolo secondario
    'equipment':               6,   # Attrezzatura principale
    'movement_pattern_1':      7,   # Schema di movimento n. 1
    'movement_pattern_2':      8,   # Schema di movimento n. 2
    'body_region':             9,   # Regione corporea
    'exercise_classification': 10,  # Classificazione primaria dell'esercizio fisico
}


def clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    return s


def unique_slug(name, used):
    base = slugify(name)[:200] or 'esercizio'
    candidate = base
    n = 2
    while candidate in used:
        suffix = f'-{n}'
        candidate = (base[:200 - len(suffix)] + suffix)
        n += 1
    used.add(candidate)
    return candidate


def main():
    xlsx_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f'ERRORE: file non trovato: {xlsx_path}')
        sys.exit(1)

    print(f'>> File:  {xlsx_path}')

    existing = Exercise.objects.count()
    if existing > 0:
        print(f'>> Tabella Exercise contiene {existing} righe.')
        ans = input("   Cancello tutto e ricarico da zero? [s/N]: ").strip().lower()
        if ans not in ('s', 'si', 'y', 'yes'):
            print('Annullato.')
            sys.exit(0)
        with transaction.atomic():
            deleted, _ = Exercise.objects.all().delete()
            print(f'>> Eliminate {deleted} righe (cascade incluso).')

    print('>> Apertura xlsx (può richiedere qualche secondo)...')
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Row 1 = empty band, Row 2 = header, Row 3+ = data.
    HEADER_ROW = 2
    total_rows = ws.max_row - HEADER_ROW
    print(f'>> Foglio: "{ws.title}", righe dati: {total_rows}')

    used_slugs = set()
    inserted = 0
    skipped = 0

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        name_cell = ws.cell(row=r, column=COLUMN_INDEX['name'] + 1)
        name = clean(name_cell.value)
        if not name:
            skipped += 1
            continue

        # Video URL: prefer hyperlink target, else cell text if it looks like a URL
        video_cell = ws.cell(row=r, column=COLUMN_INDEX['video_url'] + 1)
        video_url = ''
        if video_cell.hyperlink and video_cell.hyperlink.target:
            video_url = video_cell.hyperlink.target.strip()
        else:
            txt = clean(video_cell.value)
            if re.match(r'^https?://', txt):
                video_url = txt

        slug = unique_slug(name, used_slugs)

        Exercise.objects.create(
            name=name,
            slug=slug,
            video_url=video_url or None,
            difficulty_level=clean(ws.cell(row=r, column=COLUMN_INDEX['difficulty_level'] + 1).value) or None,
            target_muscle_group=clean(ws.cell(row=r, column=COLUMN_INDEX['target_muscle_group'] + 1).value) or None,
            primary_muscle=clean(ws.cell(row=r, column=COLUMN_INDEX['primary_muscle'] + 1).value) or None,
            secondary_muscle=clean(ws.cell(row=r, column=COLUMN_INDEX['secondary_muscle'] + 1).value) or None,
            equipment=clean(ws.cell(row=r, column=COLUMN_INDEX['equipment'] + 1).value) or None,
            movement_pattern_1=clean(ws.cell(row=r, column=COLUMN_INDEX['movement_pattern_1'] + 1).value) or None,
            movement_pattern_2=clean(ws.cell(row=r, column=COLUMN_INDEX['movement_pattern_2'] + 1).value) or None,
            body_region=clean(ws.cell(row=r, column=COLUMN_INDEX['body_region'] + 1).value) or None,
            exercise_classification=clean(ws.cell(row=r, column=COLUMN_INDEX['exercise_classification'] + 1).value) or None,
        )
        inserted += 1
        if inserted % 25 == 0 or inserted == total_rows:
            print(f'   [{inserted}/{total_rows}] {name[:60]}')

    print()
    print(f'== Fatto. Inseriti: {inserted}  |  Saltati (riga senza nome): {skipped}  |  Totale tabella: {Exercise.objects.count()}')


if __name__ == '__main__':
    main()
