#!/usr/bin/env python
"""
Standalone importer for Calcoli_Macro_copia.xlsx -> Food table.

Usage:
    cd WebApp/src
    ../../venv/bin/python import_foods.py
    # or with explicit path:
    ../../venv/bin/python import_foods.py ../../calcoli_macro.xlsx

Behavior:
- Loads xlsx (header row 1, data from row 2).
- Handles "strict" OOXML conformance by rewriting namespaces to transitional
  into a temp file before openpyxl reads it.
- If Food table has rows, prompts for confirmation, then deletes everything
  before reinserting (cascade applies to MealItem rows).
- Inserts each row, prints progress.
- After successful import, deletes the source xlsx file.

Column mapping (xlsx header -> Food.<attr> (db_column)):
    Nome_Alimento            -> nome_alimento            (Nome_Alimento)
    Categoria_Alimento       -> categoria_alimento       (Categoria_Alimento)
    Energia(Kcal)            -> energia_kcal             (Energia(Kcal))
    Proteine(g)              -> proteine_g               (Proteine(g))
    Lipidi(g)                -> lipidi_g                 (Lipidi(g))
    Colesterolo(mg)          -> colesterolo_mg           (Colesterolo(mg))
    Carboidrati(g)           -> carboidrati_g            (Carboidrati(g))
    Carboidrati_Solubili(g)  -> carboidrati_solubili_g   (Carboidrati_Solubili(g))
    Fibra(g)                 -> fibra_g                  (Fibra(g))
    Fe(mg)                   -> fe_mg                    (Fe(mg))
    Ca(mg)                   -> ca_mg                    (Ca(mg))
    Na(mg)                   -> na_mg                    (Na(mg))
    K(mg)                    -> k_mg                     (K(mg))
    P(mg)                    -> p_mg                     (P(mg))
    Zn(mg)                   -> zn_mg                    (Zn(mg))
    Mg(mg)                   -> mg_mg                    (Mg(mg))
    Cu(mg)                   -> cu_mg                    (Cu(mg))
    Se(ug)                   -> se_ug                    (Se(ug))
    I(ug)                    -> i_ug                     (I(ug))
    Mn(mg)                   -> mn_mg                    (Mn(mg))
    Vit_B1(mg)               -> vit_b1_mg                (Vit_B1(mg))
    Vit_B2(mg)               -> vit_b2_mg                (Vit_B2(mg))
    Vit_C(mg)                -> vit_c_mg                 (Vit_C(mg))
    Niacina(mg)              -> niacina_mg               (Niacina(mg))
    Vit_B6(mg)               -> vit_b6_mg                (Vit_B6(mg))
    Folati(ug)               -> folati_ug                (Folati(ug))
    Vit_B12(ug)              -> vit_b12_ug               (Vit_B12(ug))
    Lipidi_Saturi(g)         -> lipidi_saturi_g          (Lipidi_Saturi(g))
    Isoleucina(mg)           -> isoleucina_mg            (Isoleucina(mg))
    Leucina(mg)              -> leucina_mg               (Leucina(mg))
    Valina(mg)               -> valina_mg                (Valina(mg))
    Lattosio(g)              -> lattosio_g               (Lattosio(g))
"""
import os
import sys
import shutil
import tempfile
import zipfile
import warnings
import django
from pathlib import Path

warnings.filterwarnings('ignore')

# Bootstrap Django
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.db import transaction
from openpyxl import load_workbook

from domain.nutrition.models import Food


DEFAULT_XLSX = Path(__file__).resolve().parents[2] / 'calcoli_macro.xlsx'

# Python attr names on Food, in same order as xlsx columns 1..32.
# The xlsx header (Italian, with parens) maps to the db_column on each field;
# Python code talks to fields via these snake_case attrs.
FIELDS = [
    'nome_alimento',
    'categoria_alimento',
    'energia_kcal',
    'proteine_g',
    'lipidi_g',
    'colesterolo_mg',
    'carboidrati_g',
    'carboidrati_solubili_g',
    'fibra_g',
    'fe_mg',
    'ca_mg',
    'na_mg',
    'k_mg',
    'p_mg',
    'zn_mg',
    'mg_mg',
    'cu_mg',
    'se_ug',
    'i_ug',
    'mn_mg',
    'vit_b1_mg',
    'vit_b2_mg',
    'vit_c_mg',
    'niacina_mg',
    'vit_b6_mg',
    'folati_ug',
    'vit_b12_ug',
    'lipidi_saturi_g',
    'isoleucina_mg',
    'leucina_mg',
    'valina_mg',
    'lattosio_g',
]

TEXT_FIELDS = {'nome_alimento', 'categoria_alimento'}

# Strict OOXML -> Transitional namespace map (so openpyxl can read the file)
STRICT_TO_TRANS = {
    'http://purl.oclc.org/ooxml/spreadsheetml/main':
        'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'http://purl.oclc.org/ooxml/officeDocument/relationships':
        'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'http://purl.oclc.org/ooxml/drawingml/main':
        'http://schemas.openxmlformats.org/drawingml/2006/main',
    'http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing':
        'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'http://purl.oclc.org/ooxml/officeDocument/sharedTypes':
        'http://schemas.openxmlformats.org/officeDocument/2006/sharedTypes',
    'http://purl.oclc.org/ooxml/officeDocument/extendedProperties':
        'http://schemas.openxmlformats.org/officeDocument/2006/extended-properties',
    'http://purl.oclc.org/ooxml/officeDocument/customProperties':
        'http://schemas.openxmlformats.org/officeDocument/2006/custom-properties',
    'http://purl.oclc.org/ooxml/officeDocument/relationships/officeDocument':
        'http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument',
}


def normalize_strict_xlsx(src_path: Path) -> Path:
    """Rewrite a Strict OOXML xlsx into a tempfile with Transitional namespaces."""
    fd, tmp_name = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    tmp_path = Path(tmp_name)

    with zipfile.ZipFile(src_path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(('.xml', '.rels')):
                try:
                    text = data.decode('utf-8')
                except UnicodeDecodeError:
                    zout.writestr(item, data)
                    continue
                for strict, trans in STRICT_TO_TRANS.items():
                    text = text.replace(strict, trans)
                # Drop conformance="strict" attribute
                text = text.replace(' conformance="strict"', '')
                data = text.encode('utf-8')
            zout.writestr(item, data)
    return tmp_path


def clean_text(v):
    if v is None:
        return ''
    return str(v).strip()


def clean_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '.')
    if not s or s in ('-', 'tr', 'tracce', 'n.d.', 'nd', 'ND'):
        return 0.0
    # Strip anything non-numeric except leading minus and a single dot
    keep = []
    seen_dot = False
    for i, ch in enumerate(s):
        if ch.isdigit():
            keep.append(ch)
        elif ch == '.' and not seen_dot:
            keep.append(ch)
            seen_dot = True
        elif ch == '-' and i == 0:
            keep.append(ch)
    try:
        return float(''.join(keep)) if keep else 0.0
    except ValueError:
        return 0.0


def main():
    xlsx_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f'ERRORE: file non trovato: {xlsx_path}')
        sys.exit(1)

    print(f'>> File:  {xlsx_path}')

    existing = Food.objects.count()
    if existing > 0:
        print(f'>> Tabella Food contiene {existing} righe.')
        ans = input("   Cancello tutto e ricarico da zero? [s/N]: ").strip().lower()
        if ans not in ('s', 'si', 'y', 'yes'):
            print('Annullato.')
            sys.exit(0)
        with transaction.atomic():
            deleted, _ = Food.objects.all().delete()
            print(f'>> Eliminate {deleted} righe (cascade incluso).')

    print('>> Normalizzazione xlsx (strict -> transitional) ...')
    tmp_xlsx = normalize_strict_xlsx(xlsx_path)

    try:
        print('>> Apertura xlsx (puo` richiedere qualche secondo) ...')
        wb = load_workbook(tmp_xlsx, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        print(f'>> Foglio: "{ws.title}"')

        HEADER_ROW = 1
        BLANK_STOP = 50  # stop after this many consecutive nameless rows
        inserted = 0
        skipped = 0
        consecutive_blank = 0

        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            name = clean_text(row[0]) if row and len(row) > 0 else ''
            if not name:
                skipped += 1
                consecutive_blank += 1
                if consecutive_blank >= BLANK_STOP:
                    print(f'>> Stop: {BLANK_STOP} righe consecutive senza nome.')
                    break
                continue
            consecutive_blank = 0

            kwargs = {}
            for idx, field in enumerate(FIELDS):
                value = row[idx] if idx < len(row) else None
                if field in TEXT_FIELDS:
                    kwargs[field] = clean_text(value) or None if field == 'categoria_alimento' else clean_text(value)
                else:
                    kwargs[field] = clean_float(value)

            Food.objects.create(**kwargs)
            inserted += 1
            if inserted % 25 == 0:
                print(f'   [{inserted}] {name[:60]}')

        print()
        print(f'== Fatto. Inseriti: {inserted}  |  Saltati (riga senza nome): {skipped}  |  Totale tabella: {Food.objects.count()}')
    finally:
        try:
            tmp_xlsx.unlink()
        except OSError:
            pass

    # Cancella file sorgente
    try:
        xlsx_path.unlink()
        print(f'>> File sorgente eliminato: {xlsx_path}')
    except OSError as e:
        print(f'!! Impossibile eliminare il file sorgente: {e}')


if __name__ == '__main__':
    main()

